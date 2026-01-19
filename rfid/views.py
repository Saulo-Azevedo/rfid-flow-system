# rfid/views.py

import logging

logger = logging.getLogger(__name__)

import json  # <--- Necessário para ler o corpo da requisição
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.mail import EmailMessage
from django.db.models import Count, Q

# ... outros imports que já existiam ...
from django.http import JsonResponse  # <--- Necessário para a API
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt  # <--- O QUE FALTOU
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .forms import BotijaoForm
from .models import Botijao, LeituraRFID, LogAuditoria

# -----------------------
# Helpers de requalificação
# -----------------------


def _classificar_requal(botijao, hoje):
    """
    Retorna um rótulo simples para o status da requalificação:
    vencida / proxima / em_dia / sem_data
    """
    proxima = botijao.data_proxima_requalificacao
    if not proxima:
        return "sem_data"
    if proxima <= hoje:
        return "vencida"
    if hoje < proxima <= hoje + timedelta(days=90):
        return "proxima"
    return "em_dia"


# -----------------------
# Dashboard
# -----------------------


@login_required
def dashboard(request):
    hoje = timezone.now().date()

    # Estatísticas principais
    total_botijoes = Botijao.objects.filter(deletado=False).count()
    leituras_hoje = LeituraRFID.objects.filter(data_hora__date=hoje).count()

    # Botijões ativos = próximos da requalificação ou dentro da validade
    botijoes_ativos = (
        Botijao.objects.filter(deletado=False)
        .exclude(status_requalificacao="vencida")
        .count()
    )

    # --- Últimos botijões cadastrados + total de leituras ---

    RFID_TAG_REGEX = r"^[0-9A-Fa-f]{24}$|^[0-9A-Fa-f]{32}$"

    botijoes = (
        Botijao.objects.filter(deletado=False, tag_rfid__regex=RFID_TAG_REGEX)
        .annotate(num_leituras=Count("leituras"))
        .order_by("-id")[:10]
    )

    # --- Leituras dos últimos 7 dias (gráfico e estatísticas futuras) ---
    leituras_7_dias = []
    for i in range(6, -1, -1):
        dia = hoje - timedelta(days=i)
        leituras_7_dias.append(
            {
                "data": dia.strftime("%d/%m"),
                "total": LeituraRFID.objects.filter(data_hora__date=dia).count(),
            }
        )

    # Requalificação
    cilindros = Botijao.objects.filter(deletado=False)
    requal_vencidas = []
    requal_proximas = []
    requal_em_dia = []
    requal_sem_data = []

    for c in cilindros:
        status = _classificar_requal(c, hoje)
        if status == "vencida":
            requal_vencidas.append(c)
        elif status == "proxima":
            requal_proximas.append(c)
        elif status == "em_dia":
            requal_em_dia.append(c)
        else:
            requal_sem_data.append(c)

    # Últimas leituras
    ultimas_leituras = (
        LeituraRFID.objects.select_related("botijao")
        .filter(botijao__tag_rfid__regex=RFID_TAG_REGEX)
        .order_by("-data_hora")[:10]
    )

    # CONTEXTO FINAL — **somente UM return**, no fim!
    context = {
        "total_botijoes": total_botijoes,
        "leituras_hoje": leituras_hoje,
        "botijoes_ativos": botijoes_ativos,
        "botijoes": botijoes,
        "ultimas_leituras": ultimas_leituras,
        "leituras_7_dias": leituras_7_dias,
        "qtd_requal_vencidas": len(requal_vencidas),
        "qtd_requal_proximas": len(requal_proximas),
        "qtd_requal_em_dia": len(requal_em_dia),
        "qtd_requal_sem_data": len(requal_sem_data),
    }

    return render(request, "rfid/dashboard.html", context)


@login_required
def dashboard_api(request):
    """Versão JSON do dashboard para uso com AJAX, se necessário."""
    hoje = timezone.now().date()

    total_cilindros = Botijao.objects.filter(deletado=False).count()
    total_leituras_hoje = LeituraRFID.objects.filter(data_hora__date=hoje).count()

    leituras_7_dias = []
    for i in range(6, -1, -1):
        dia = hoje - timedelta(days=i)
        count = LeituraRFID.objects.filter(data_hora__date=dia).count()
        leituras_7_dias.append(
            {
                "data": dia.strftime("%d/%m"),
                "total": count,
            }
        )

    cilindros = list(Botijao.objects.filter(deletado=False))
    requal_vencidas = []
    requal_proximas = []
    requal_em_dia = []
    requal_sem_data = []

    for c in cilindros:
        status = _classificar_requal(c, hoje)
        if status == "vencida":
            requal_vencidas.append(c)
        elif status == "proxima":
            requal_proximas.append(c)
        elif status == "em_dia":
            requal_em_dia.append(c)
        else:
            requal_sem_data.append(c)

    requal_proximas_ordenadas = sorted(
        requal_proximas, key=lambda x: x.data_proxima_requalificacao or hoje
    )[:10]

    proximas_data = [
        {
            "tag_rfid": c.tag_rfid,
            "fabricante": c.fabricante or "-",
            "numero_serie": c.numero_serie or "-",
            "data_ultima_requalificacao": (
                c.data_ultima_requalificacao.strftime("%d/%m/%Y")
                if c.data_ultima_requalificacao
                else "-"
            ),
            "data_proxima_requalificacao": (
                c.data_proxima_requalificacao.strftime("%d/%m/%Y")
                if c.data_proxima_requalificacao
                else "-"
            ),
        }
        for c in requal_proximas_ordenadas
    ]

    return JsonResponse(
        {
            "total_cilindros": total_cilindros,
            "total_leituras_hoje": total_leituras_hoje,
            "leituras_7_dias": leituras_7_dias,
            "qtd_requal_vencidas": len(requal_vencidas),
            "qtd_requal_proximas": len(requal_proximas),
            "qtd_requal_em_dia": len(requal_em_dia),
            "qtd_requal_sem_data": len(requal_sem_data),
            "requal_proximas": proximas_data,
        }
    )


# -----------------------
# Cadastro / edição de Botijão
# -----------------------


@login_required
def novo_botijao(request):
    if request.method == "POST":
        form = BotijaoForm(request.POST)
        if form.is_valid():
            botijao = form.save()
            messages.success(
                request, f"Botijão {botijao.tag_rfid} cadastrado com sucesso."
            )
            return redirect("dashboard")
    else:
        form = BotijaoForm()

    return render(request, "botijao_form.html", {"form": form})


@login_required
def editar_botijao(request, botijao_id):
    botijao = get_object_or_404(Botijao, pk=botijao_id, deletado=False)

    if request.method == "POST":
        form = BotijaoForm(request.POST, instance=botijao)
        if form.is_valid():
            form.save()
            messages.success(
                request, f"Botijão {botijao.tag_rfid} atualizado com sucesso."
            )
            return redirect("historico_botijao", botijao_id=botijao.id)
    else:
        form = BotijaoForm(instance=botijao)

    return render(request, "botijao_form.html", {"form": form, "botijao": botijao})


# -----------------------
# Leitura RFID manual
# -----------------------


@login_required
def nova_leitura(request):
    if request.method == "POST":
        tag_rfid = request.POST.get("tag_rfid", "").strip()
        operador = request.POST.get("operador", "").strip()
        observacao = request.POST.get("observacao", "").strip()

        if not tag_rfid:
            messages.error(request, "Tag RFID é obrigatória.")
            return redirect("nova_leitura")

        botijao, criado = Botijao.objects.get_or_create(tag_rfid=tag_rfid)

        leitura = LeituraRFID.objects.create(
            botijao=botijao,
            operador=operador or None,
            observacao=observacao or None,
        )

        try:
            LogAuditoria.criar_log(
                botijao=botijao,
                acao="leitura",
                usuario=request.user if request.user.is_authenticated else None,
                descricao=f"Leitura manual registrada. Observação: {observacao or '-'}",
                dados_anteriores=None,
                dados_novos={"leitura_id": leitura.id},
            )
        except Exception:
            pass

        if criado:
            messages.success(
                request,
                f"Novo botijão cadastrado e leitura registrada. Tag: {tag_rfid}",
            )
        else:
            messages.success(request, f"Leitura registrada. Tag: {tag_rfid}")

        return redirect("dashboard")

    return render(request, "nova_leitura.html")


# -----------------------
# Relatórios de requalificação e leituras
# -----------------------

RFID_TAG_REGEX = r"^[0-9A-Fa-f]{24}$|^[0-9A-Fa-f]{32}$"


@login_required
def relatorios(request):
    status = request.GET.get("status", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")
    data_tipo = (
        request.GET.get("data_tipo") or "cadastro"
    ).strip()  # "cadastro" | "leitura"
    tipo = request.GET.get("tipo", "")  # "" | "rfid" | "barcode"

    botijoes = Botijao.objects.filter(deletado=False).annotate(
        num_leituras=Count("leituras")
    )

    # FILTRO POR STATUS
    if status:
        botijoes = botijoes.filter(status=status)

    # FILTROS POR DATA (CADASTRO OU LEITURA)
    if data_inicio or data_fim:
        if data_tipo == "leitura":
            if data_inicio:
                botijoes = botijoes.filter(leituras__data_hora__date__gte=data_inicio)
            if data_fim:
                botijoes = botijoes.filter(leituras__data_hora__date__lte=data_fim)
            botijoes = botijoes.distinct()
        else:
            # cadastro (padrão antigo)
            if data_inicio:
                botijoes = botijoes.filter(data_cadastro__date__gte=data_inicio)
            if data_fim:
                botijoes = botijoes.filter(data_cadastro__date__lte=data_fim)

    # FILTRO POR TIPO (RFID x BARCODE)
    if tipo == "rfid":
        botijoes = botijoes.filter(tag_rfid__regex=RFID_TAG_REGEX)
    elif tipo == "barcode":
        botijoes = botijoes.exclude(tag_rfid__regex=RFID_TAG_REGEX)

    botijoes = botijoes.order_by("-data_cadastro")

    # TOTAL DE LEITURAS SOMADAS (mantive igual seu padrão)
    total_leituras = sum([b.num_leituras for b in botijoes])

    context = {
        "botijoes": botijoes,
        "total_filtrado": botijoes.count(),
        "total_leituras": total_leituras,
        "status_choices": Botijao.STATUS_CHOICES,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "data_tipo": data_tipo,
        "status_selected": status,
        "tipo_selected": tipo,  # mantém o select marcado no template
    }

    return render(request, "rfid/relatorios.html", context)

    # ✅ NOVO: FILTRO POR TIPO (RFID x BARCODE)
    if tipo == "rfid":
        botijoes = botijoes.filter(tag_rfid__regex=RFID_TAG_REGEX)
    elif tipo == "barcode":
        botijoes = botijoes.exclude(tag_rfid__regex=RFID_TAG_REGEX)

    botijoes = botijoes.order_by("-data_cadastro")

    # TOTAL DE LEITURAS SOMADAS
    total_leituras = sum([b.num_leituras for b in botijoes])

    context = {
        "botijoes": botijoes,
        "total_filtrado": botijoes.count(),
        "total_leituras": total_leituras,
        "status_choices": Botijao.STATUS_CHOICES,
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "data_tipo": data_tipo,
        "status_selected": status,
        # ✅ NOVO: pra manter o select marcado no template
        "tipo_selected": tipo,
    }

    return render(request, "rfid/relatorios.html", context)


@login_required
def relatorios_api(request):
    status = request.GET.get("status", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")

    qs = Botijao.objects.filter(deletado=False).annotate(num_leituras=Count("leituras"))

    if status:
        qs = qs.filter(status=status)

    if data_inicio:
        qs = qs.filter(data_cadastro__date__gte=data_inicio)

    if data_fim:
        qs = qs.filter(data_cadastro__date__lte=data_fim)

    qs = qs.order_by("-data_cadastro")

    botijoes_data = []
    for b in qs:
        botijoes_data.append(
            {
                "tag_rfid": b.tag_rfid,
                "numero_serie": b.numero_serie or "-",
                "fabricante": b.fabricante or "-",
                "status_display": b.get_status_display(),
                "status_requalificacao_display": b.get_status_requalificacao_display(),
                "total_leituras": b.num_leituras,
                "data_cadastro": b.data_cadastro.strftime("%d/%m/%Y %H:%M"),
            }
        )

    return JsonResponse({"botijoes": botijoes_data, "total_filtrado": qs.count()})


# -----------------------
# Exportar Excel
# -----------------------


@login_required
def exportar_excel(request):
    hoje = timezone.now().date()

    # ✅ mesmos filtros do relatório
    data_tipo = (request.GET.get("data_tipo") or "cadastro").strip()
    status = request.GET.get("status", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")
    tipo = request.GET.get("tipo", "")  # "", "rfid", "barcode"

    # queryset base
    qs = Botijao.objects.filter(deletado=False).annotate(num_leituras=Count("leituras"))

    # status
    if status:
        qs = qs.filter(status=status)

    # FILTRO POR DATA (CADASTRO OU LEITURA)
    if data_inicio or data_fim:
        if data_tipo == "leitura":
            if data_inicio:
                qs = qs.filter(leituras__data_hora__date__gte=data_inicio)
            if data_fim:
                qs = qs.filter(leituras__data_hora__date__lte=data_fim)
            qs = qs.distinct()
        else:
            # cadastro (padrão antigo)
            if data_inicio:
                qs = qs.filter(data_cadastro__date__gte=data_inicio)
            if data_fim:
                qs = qs.filter(data_cadastro__date__lte=data_fim)

    # tipo (RFID x Barcode)
    if tipo == "rfid":
        qs = qs.filter(tag_rfid__regex=RFID_TAG_REGEX)
    elif tipo == "barcode":
        qs = qs.exclude(tag_rfid__regex=RFID_TAG_REGEX)

    qs = qs.order_by("tag_rfid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cilindros RFID"

    headers = [
        "Tag RFID",
        "Fabricante",
        "Número de Série",
        "Tara (kg)",
        "Última requalificação",
        "Próxima requalificação",
        "Dias para requalificação",
        "Status requalificação",
        "Penúltima envasadora",
        "Data penúltimo envasamento",
        "Última envasadora",
        "Data último envasamento",
    ]
    ws.append(headers)

    header_fill = PatternFill(
        start_color="00D4FF", end_color="00D4FF", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for c in qs:
        status_requal = _classificar_requal(c, hoje)
        dias = None
        if c.data_proxima_requalificacao:
            dias = (c.data_proxima_requalificacao - hoje).days

        ws.append(
            [
                c.tag_rfid,
                c.fabricante or "-",
                c.numero_serie or "-",
                float(c.tara) if c.tara is not None else "-",
                (
                    c.data_ultima_requalificacao.strftime("%d/%m/%Y")
                    if c.data_ultima_requalificacao
                    else "-"
                ),
                (
                    c.data_proxima_requalificacao.strftime("%d/%m/%Y")
                    if c.data_proxima_requalificacao
                    else "-"
                ),
                dias if dias is not None else "-",
                status_requal,
                c.penultima_envasadora or "-",
                (
                    c.data_penultimo_envasamento.strftime("%d/%m/%Y")
                    if c.data_penultimo_envasamento
                    else "-"
                ),
                c.ultima_envasadora or "-",
                (
                    c.data_ultimo_envasamento.strftime("%d/%m/%Y")
                    if c.data_ultimo_envasamento
                    else "-"
                ),
            ]
        )

    # auto width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"relatorio_cilindros_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -----------------------
# Histórico por botijão
# -----------------------


@login_required
def historico_botijao(request, botijao_id):
    botijao = get_object_or_404(Botijao, id=botijao_id, deletado=False)
    leituras = LeituraRFID.objects.filter(botijao=botijao).order_by("-data_hora")
    context = {
        "botijao": botijao,
        "leituras": leituras,
    }
    return render(request, "historico_botijao.html", context)


@login_required
def buscar_historico(request):
    query = request.GET.get("q", "").strip()
    status = request.GET.get("status", "")
    status_requal = request.GET.get("status_requal", "")
    data_req_inicio = request.GET.get("data_req_inicio", "")
    data_req_fim = request.GET.get("data_req_fim", "")
    data_env_inicio = request.GET.get("data_env_inicio", "")
    data_env_fim = request.GET.get("data_env_fim", "")
    operador = request.GET.get("operador", "").strip()

    resultados = []

    qs = (
        Botijao.objects.filter(deletado=False)
        .annotate(num_leituras=Count("leituras"))
        .prefetch_related("leituras")
    )

    # -------------------------
    # 🔍 Filtros Avançados
    # -------------------------

    if query:
        qs = qs.filter(
            Q(tag_rfid__icontains=query)
            | Q(numero_serie__icontains=query)
            | Q(fabricante__icontains=query)
        )

    if status:
        qs = qs.filter(status=status)

    if status_requal:
        qs = qs.filter(status_requalificacao=status_requal)

    if data_req_inicio:
        qs = qs.filter(data_ultima_requalificacao__gte=data_req_inicio)

    if data_req_fim:
        qs = qs.filter(data_ultima_requalificacao__lte=data_req_fim)

    if data_env_inicio:
        qs = qs.filter(data_ultimo_envasamento__gte=data_env_inicio)

    if data_env_fim:
        qs = qs.filter(data_ultimo_envasamento__lte=data_env_fim)

    # -------------------------
    # 🔍 Filtrar por OPERADOR
    # -------------------------
    if operador:
        qs = qs.filter(leituras__operador__icontains=operador).distinct()

    # -------------------------
    # 🔍 Montagem dos Resultados
    # -------------------------

    for b in qs:
        leituras = list(b.leituras.order_by("-data_hora")[:10])

        leituras_data = [
            {
                "data_hora": l.data_hora.strftime("%d/%m/%Y %H:%M:%S"),
                "operador": l.operador or "-",
                "observacao": l.observacao or "-",
            }
            for l in leituras
        ]

        resultados.append(
            {
                "botijao": {
                    "id": b.id,
                    "tag_rfid": b.tag_rfid,
                    "numero_serie": b.numero_serie or "-",
                    "fabricante": b.fabricante or "-",
                    "tara": b.tara or "-",
                    "data_ultima_requalificacao": b.data_ultima_requalificacao,
                    "data_proxima_requalificacao": b.data_proxima_requalificacao,
                    "status_requalificacao_display": b.get_status_requalificacao_display(),
                    "penultima_envasadora": b.penultima_envasadora or "-",
                    "data_penultimo_envasamento": b.data_penultimo_envasamento,
                    "ultima_envasadora": b.ultima_envasadora or "-",
                    "data_ultimo_envasamento": b.data_ultimo_envasamento,
                    "status": b.get_status_display(),
                },
                "leituras": leituras_data,
                "total_leituras": b.num_leituras,
            }
        )

    context = {
        "query": query,
        "resultados": resultados,
        "total_encontrados": len(resultados),
        "status_choices": Botijao.STATUS_CHOICES,
        "status_requal_choices": Botijao.STATUS_REQUALIFICACAO_CHOICES,
    }

    return render(request, "historico_busca.html", context)


# -----------------------
# Enviar relatório por e-mail
# -----------------------


@login_required
def enviar_email_view(request):
    # ✅ Captura filtros (GET quando vem de Relatórios, POST quando envia o form)
    data_tipo = (
        request.POST.get("data_tipo") or request.GET.get("data_tipo") or "cadastro"
    ).strip()

    tipo = (
        request.POST.get("tipo") or request.GET.get("tipo") or ""
    ).strip()  # "", "rfid", "barcode"
    status_filtro = (
        request.POST.get("status") or request.GET.get("status") or ""
    ).strip()
    data_inicio = (
        request.POST.get("data_inicio") or request.GET.get("data_inicio") or ""
    ).strip()
    data_fim = (
        request.POST.get("data_fim") or request.GET.get("data_fim") or ""
    ).strip()

    # ✅ Resumo humano dos filtros (para mostrar no template e também no corpo do e-mail)
    if tipo == "rfid":
        tipo_label = "Somente RFID"
    elif tipo == "barcode":
        tipo_label = "Somente Código de Barras"
    else:
        tipo_label = "Todos"

    status_label = status_filtro if status_filtro else "Todos"

    if data_inicio or data_fim:
        periodo_label = f"{data_inicio or '—'} até {data_fim or '—'}"
    else:
        periodo_label = "Todos"

    data_tipo_label = (
        "Data da Leitura" if data_tipo == "leitura" else "Data de Cadastro"
    )

    filtro_resumo = (
        f"Tipo: {tipo_label} | "
        f"Status: {status_label} | "
        f"Data: {data_tipo_label} | "
        f"Período: {periodo_label}"
    )

    # Helper: renderiza o template SEM perder filtros
    def render_pagina(destinatario_value=""):
        return render(
            request,
            "enviar_email.html",
            {
                "tipo_selected": tipo,
                "status_selected": status_filtro,
                "data_inicio": data_inicio,
                "data_fim": data_fim,
                "filtro_resumo": filtro_resumo,
                # ✅ para manter o campo preenchido caso dê erro
                "destinatario_value": destinatario_value,
            },
        )

    # ✅ Se for GET: só renderiza a página já mostrando o resumo e preservando filtros
    if request.method != "POST":
        return render_pagina()

    # ---------------------------
    # POST: enviar e-mail
    # ---------------------------
    destinatario = request.POST.get("destinatario", "").strip()

    if not destinatario:
        messages.error(request, "Email de destinatário é obrigatório.")
        # ✅ não redireciona (senão perde filtros)
        return render_pagina(destinatario_value=destinatario)

    try:
        from io import BytesIO

        hoje = timezone.now().date()

        # ✅ Base queryset (igual exportar_excel)
        qs = Botijao.objects.filter(deletado=False).annotate(
            num_leituras=Count("leituras")
        )

        # filtro status
        if status_filtro:
            qs = qs.filter(status=status_filtro)

        # FILTRO POR DATA (CADASTRO OU LEITURA)
        if data_inicio or data_fim:
            if data_tipo == "leitura":
                if data_inicio:
                    qs = qs.filter(leituras__data_hora__date__gte=data_inicio)
                if data_fim:
                    qs = qs.filter(leituras__data_hora__date__lte=data_fim)
                qs = qs.distinct()
            else:
                if data_inicio:
                    qs = qs.filter(data_cadastro__date__gte=data_inicio)
                if data_fim:
                    qs = qs.filter(data_cadastro__date__lte=data_fim)

        # filtro tipo (rfid x barcode)
        if tipo == "rfid":
            qs = qs.filter(tag_rfid__regex=RFID_TAG_REGEX)
        elif tipo == "barcode":
            qs = qs.exclude(tag_rfid__regex=RFID_TAG_REGEX)

        qs = qs.order_by("tag_rfid")

        # ---------------------------
        # ✅ Gera Excel em memória
        # ---------------------------
        wb = Workbook()
        ws = wb.active

        header_fill = PatternFill(
            start_color="00D4FF", end_color="00D4FF", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # ✅ Se tipo=barcode: gera planilha enxuta "de outra forma"
        if tipo == "barcode":
            ws.title = "Códigos de Barras"
            headers = [
                "Código de Barras",
                "Status",
                "Total Leituras",
                "Data Cadastro",
                "Observação",
            ]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for b in qs:
                ws.append(
                    [
                        b.tag_rfid,
                        b.get_status_display(),
                        b.num_leituras,
                        (
                            b.data_cadastro.strftime("%d/%m/%Y %H:%M")
                            if b.data_cadastro
                            else "-"
                        ),
                        getattr(b, "observacao", None)
                        or getattr(b, "observacao_interna", None)
                        or "-",
                    ]
                )

            filename = (
                f"relatorio_barcode_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

        else:
            # RFID ou TODOS: mesma estrutura do seu exportar_excel atual
            ws.title = "Cilindros RFID"

            headers = [
                "Tag RFID",
                "Fabricante",
                "Número de Série",
                "Tara (kg)",
                "Última requalificação",
                "Próxima requalificação",
                "Dias para requalificação",
                "Status requalificação",
                "Penúltima envasadora",
                "Data penúltimo envasamento",
                "Última envasadora",
                "Data último envasamento",
            ]
            ws.append(headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for c in qs:
                status_requal = _classificar_requal(c, hoje)
                dias = None
                if c.data_proxima_requalificacao:
                    dias = (c.data_proxima_requalificacao - hoje).days

                ws.append(
                    [
                        c.tag_rfid,
                        c.fabricante or "-",
                        c.numero_serie or "-",
                        float(c.tara) if c.tara is not None else "-",
                        (
                            c.data_ultima_requalificacao.strftime("%d/%m/%Y")
                            if c.data_ultima_requalificacao
                            else "-"
                        ),
                        (
                            c.data_proxima_requalificacao.strftime("%d/%m/%Y")
                            if c.data_proxima_requalificacao
                            else "-"
                        ),
                        dias if dias is not None else "-",
                        status_requal,
                        c.penultima_envasadora or "-",
                        (
                            c.data_penultimo_envasamento.strftime("%d/%m/%Y")
                            if c.data_penultimo_envasamento
                            else "-"
                        ),
                        c.ultima_envasadora or "-",
                        (
                            c.data_ultimo_envasamento.strftime("%d/%m/%Y")
                            if c.data_ultimo_envasamento
                            else "-"
                        ),
                    ]
                )

            filename = (
                f"relatorio_cilindros_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

        # auto width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # ---------------------------
        # ✅ Corpo do e-mail
        # ---------------------------
        data_hora_str = timezone.now().strftime("%d/%m/%Y às %H:%M")
        total_cilindros = qs.count()

        # Mantém seu total de leituras global (você já fazia assim)
        total_leituras = LeituraRFID.objects.count()

        corpo_html = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
            <meta charset="UTF-8">
        </head>
        <body>
            <div style="font-family:Segoe UI, Tahoma, Geneva, Verdana, sans-serif;">
                <h2>RFID FLOW - Relatório</h2>
                <p>Relatório gerado em <strong>{data_hora_str}</strong>.</p>
                <p><strong>Filtros aplicados:</strong> {filtro_resumo}</p>

                <p><strong>Total de itens no relatório:</strong> {total_cilindros}</p>
                <p><strong>Total de leituras RFID no sistema:</strong> {total_leituras}</p>

                <p>Segue anexo o arquivo Excel.</p>
            </div>
        </body>
        </html>
        """

        email = EmailMessage(
            subject=f"Relatório RFID Flow - {data_hora_str}",
            body=corpo_html,
            from_email="RFID Flow <noreply@rfidflow.com>",
            to=[destinatario],
        )
        email.content_subtype = "html"
        email.attach(
            filename,
            excel_buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        email.send(fail_silently=False)

        messages.success(request, f"Relatório enviado para {destinatario}.")
        return redirect("relatorios")

    except Exception as e:
        logger.exception(
            "Erro ao enviar email"
        )  # ✅ isso imprime traceback completo no console
        messages.error(request, f"Erro ao enviar email: {str(e)}")
        return render_pagina(destinatario_value=destinatario)


# -----------------------
# Criar admin temporário (Railway)
# -----------------------

# def criar_admin_temp(request):
#     try:
#         if not User.objects.filter(username="rfidadmin").exists():
#             User.objects.create_superuser(
#                 username="rfidadmin",
#                 email="admin@rfidflow.com",
#                 password="RFID@Admin2024!",
#             )
#             return HttpResponse(
#                 "Superusuário criado. Username: rfidadmin | Senha: RFID@Admin2024!"
#             )
#         return HttpResponse("Superusuário já existe.")
#     except Exception as e:
#         return HttpResponse(f"Erro: {str(e)}")


# Aliases para URLs antigas
historico_busca = buscar_historico
enviar_relatorio_view = enviar_email_view


# -----------------------
# API para registrar leitura RFID
# -----------------------


@csrf_exempt  # <--- Isso permite que o Android envie dados sem token de navegador
def api_registrar_leitura(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Use POST"}, status=405)

    try:
        # 1. Ler o JSON que vem do Android
        data = json.loads(request.body)

        # 2. Pegar os dados usando os nomes exatos
        tag_rfid = data.get("tag_rfid", "").strip()
        operador = data.get("operador", "PDA_C72").strip()  # Valor padrão se vier vazio
        observacao = data.get("observacao", "Leitura Mobile").strip()

        if not tag_rfid:
            return JsonResponse(
                {"success": False, "error": "Tag RFID faltando"}, status=400
            )

        # 3. Lógica do Botijão (Mantida igual a sua)
        botijao, criado = Botijao.objects.get_or_create(tag_rfid=tag_rfid)

        leitura = LeituraRFID.objects.create(
            botijao=botijao,
            operador=operador,
            observacao=observacao,
        )

        # 4. Log (O usuário será None pois não tem sessão, isso evita o crash)
        try:
            LogAuditoria.criar_log(
                botijao=botijao,
                acao="leitura",
                usuario=None,  # Android sem login envia como None ou Sistema
                descricao=f"Leitura API. Op: {operador}",
                dados_anteriores=None,
                dados_novos={"leitura_id": leitura.id},
            )
        except Exception:
            pass

        return JsonResponse(
            {"success": True, "message": "Sucesso", "id_leitura": leitura.id}
        )

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
