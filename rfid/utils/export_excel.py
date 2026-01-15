import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
from django.conf import settings
import os

# ✅ Mesmo padrão que você usou no dashboard: EPC RFID em HEX longo
RFID_TAG_REGEX = r"^[0-9A-Fa-f]{24}$|^[0-9A-Fa-f]{32}$"


def _apply_column_widths(ws, column_widths):
    """
    Aplica larguras fixas por coluna (se fornecido).
    Mantém comportamento seguro e não quebra nada.
    """
    if not column_widths:
        return
    for idx, w in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w


def _aplicar_filtro_tipo(qs, tipo, field_name):
    """
    tipo:
      - "" / None -> não filtra (exporta tudo)
      - "rfid" -> somente EPC RFID (hex longo)
      - "barcode" -> tudo que NÃO for EPC RFID (códigos de barras)
    field_name: nome do campo onde está o "código" (ex: "tag_rfid")
    """
    if not tipo:
        return qs

    lookup = f"{field_name}__regex"

    if tipo == "rfid":
        return qs.filter(**{lookup: RFID_TAG_REGEX})
    elif tipo == "barcode":
        return qs.exclude(**{lookup: RFID_TAG_REGEX})

    # Qualquer valor inesperado: não filtra (modo seguro)
    return qs


def gerar_excel_botijoes(filepath=None, status=None, data_inicio=None, data_fim=None, tipo=None):
    """
    Gera planilha Excel com botijões.

    ✅ Backward compatible:
      - chamadas antigas continuam funcionando (mesmos defaults)
    ✅ Novos filtros opcionais:
      - status: filtra Botijao.status
      - data_inicio / data_fim: filtra Botijao.data_cadastro (intervalo)
      - tipo: "", "rfid", "barcode" (baseado em tag_rfid)
    """
    from rfid.models import Botijao

    if not filepath:
        # Cria diretório temporário se não existir
        temp_dir = os.path.join(settings.BASE_DIR, "temp_exports")
        os.makedirs(temp_dir, exist_ok=True)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(temp_dir, f"relatorio_botijoes_{timestamp}.xlsx")

    # Cria workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Botijões"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cabeçalhos (mantidos)
    headers = [
        "Tag RFID",
        "Nº Série",
        "Cliente",
        "Localização",
        "Status",
        "Total Leituras",
        "Data Cadastro",
        "Última Leitura",
        "Capacidade",
        "Observações",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Dados (mesma base de antes)
    botijoes = Botijao.objects.filter(deletado=False)

    # ✅ filtros novos (sem quebrar)
    if status:
        botijoes = botijoes.filter(status=status)

    if data_inicio:
        # aceita string YYYY-MM-DD ou datetime (Django lida bem)
        botijoes = botijoes.filter(data_cadastro__date__gte=data_inicio)

    if data_fim:
        botijoes = botijoes.filter(data_cadastro__date__lte=data_fim)

    # ✅ filtro tipo (rfid / barcode / todos)
    botijoes = _aplicar_filtro_tipo(botijoes, tipo, "tag_rfid")

    # ordenação mantida
    botijoes = botijoes.order_by("-data_cadastro")

    for row, botijao in enumerate(botijoes, 2):
        ws.cell(row, 1, botijao.tag_rfid).border = border
        ws.cell(row, 2, botijao.numero_serie or "-").border = border
        ws.cell(row, 3, botijao.cliente or "-").border = border
        ws.cell(row, 4, botijao.localizacao or "-").border = border
        ws.cell(row, 5, botijao.get_status_display()).border = border

        # mantém a contagem do jeito antigo (sem mexer em models/views)
        ws.cell(row, 6, botijao.leituras.count()).border = border

        if botijao.data_cadastro:
            ws.cell(row, 7, botijao.data_cadastro.strftime("%d/%m/%Y %H:%M")).border = border
        else:
            ws.cell(row, 7, "-").border = border

        if botijao.ultima_leitura:
            ws.cell(row, 8, botijao.ultima_leitura.strftime("%d/%m/%Y %H:%M")).border = border
        else:
            ws.cell(row, 8, "-").border = border

        ws.cell(row, 9, str(botijao.capacidade)).border = border
        ws.cell(row, 10, botijao.observacao or "-").border = border

    # Ajusta largura das colunas (antes você definia mas não aplicava; aplicar não quebra)
    column_widths = [25, 20, 25, 25, 12, 15, 18, 18, 12, 40]
    _apply_column_widths(ws, column_widths)

    # Congela primeira linha
    ws.freeze_panes = "A2"

    # Salva
    wb.save(filepath)
    print(f"✅ Excel de botijões gerado: {filepath}")
    return filepath


def gerar_excel_leituras(data_inicio=None, data_fim=None, filepath=None, tipo=None):
    """
    Gera planilha Excel com histórico de leituras.

    ✅ Backward compatible:
      - assinatura antiga segue funcionando
    ✅ Novos filtros opcionais:
      - data_inicio / data_fim: filtra Leitura.data_hora (intervalo)
      - tipo: "", "rfid", "barcode" (baseado em leitura.tag_rfid)
    """
    from rfid.models import Leitura

    if not filepath:
        temp_dir = os.path.join(settings.BASE_DIR, "temp_exports")
        os.makedirs(temp_dir, exist_ok=True)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filepath = os.path.join(temp_dir, f"relatorio_leituras_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leituras"

    # Estilos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Cabeçalhos (mantidos)
    headers = ["Data/Hora", "Tag RFID", "Nº Série", "Cliente", "Operador", "Localização", "Observação"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Filtra leituras
    leituras = Leitura.objects.select_related("botijao").all()

    if data_inicio:
        leituras = leituras.filter(data_hora__gte=data_inicio)
    if data_fim:
        leituras = leituras.filter(data_hora__lte=data_fim)

    # ✅ filtro tipo (rfid / barcode / todos) baseado em leitura.tag_rfid
    leituras = _aplicar_filtro_tipo(leituras, tipo, "tag_rfid")

    # Dados
    for row, leitura in enumerate(leituras, 2):
        ws.cell(row, 1, leitura.data_hora.strftime("%d/%m/%Y %H:%M:%S")).border = border
        ws.cell(row, 2, leitura.tag_rfid).border = border
        ws.cell(
            row,
            3,
            leitura.botijao.numero_serie if leitura.botijao and leitura.botijao.numero_serie else "-",
        ).border = border
        ws.cell(
            row,
            4,
            leitura.botijao.cliente if leitura.botijao and leitura.botijao.cliente else "-",
        ).border = border
        ws.cell(row, 5, leitura.operador or "-").border = border
        ws.cell(row, 6, leitura.localizacao or "-").border = border
        ws.cell(row, 7, leitura.observacao or "-").border = border

    # Ajusta larguras (antes definia e não aplicava; aplicar não quebra)
    column_widths = [20, 25, 20, 25, 20, 25, 40]
    _apply_column_widths(ws, column_widths)

    ws.freeze_panes = "A2"

    wb.save(filepath)
    print(f"✅ Excel de leituras gerado: {filepath}")
    return filepath
